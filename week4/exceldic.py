from openpyxl import Workbook

week_teperatures = {
    'Monday': 54,
    'Tuesday': 60,
    'Wednesday': 62,
    'Thursday': 57,
    'Friday': 71,
}

workbook = Workbook()
worksheet = workbook.active
worksheet.cell(1, 1, 'Day')
worksheet.cell(1, 2, 'Temperature(F)')
worksheet.title = "Daily Temperatures"

for index, day in enumerate(week_teperatures):
    row = index + 2
    temp = week_teperatures[day]
    worksheet.cell(row, 1, day)
    worksheet.cell(row, 2, temp)

workbook.save('tempatures.xlsx')