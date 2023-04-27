from openpyxl import Workbook

favorite_foods = ['Pizza', 'Chocolate', 'Broccoli']
favorite_colors = ['Blue', 'Purple', 'Green', 'Orange']

workbook = Workbook()

worksheet = workbook.active

worksheet.title = 'Favorite Things'

worksheet.cell(1, 1, 'Favorite Foods')

worksheet.cell(1, 2, 'Favorite Colors')

for index, food in enumerate(favorite_foods):
    worksheet.cell(index + 2, 1, food)

for index, color in enumerate(favorite_colors):
    worksheet.cell(index + 2, 2, color)

workbook.save('favorites.xlsx')